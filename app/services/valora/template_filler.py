"""Fill the administrator's Valora workbook without rebuilding its layout."""

from io import BytesIO
import re

from openpyxl import load_workbook
SHEET_NAME = "Plantilla Usuario"
FIRST_YEAR_CELL = "C7"
LAST_YEAR_CELL = "C8"

# Rows containing source values. Totals and calculated rows deliberately omitted.
BALANCE_ROWS = {
    "Efectivo y Equivalentes al Efectivo": 11,
    "Cuentas por Cobrar Comerciales": 12,
    "Cuentas por Cobrar a Entidades Relacionadas": 13,
    "Otras Cuentas por Cobrar": 14,
    "Inventarios": 15,
    "Otros activos no financieros": 16,
    "Cuentas por Cobrar Comerciales y Otras Cuentas por Cobrar": 18,
    "Inversiones Financieras e inmobiliarias": 19,
    "Propiedades, Planta y Equipo": 20,
    "Depreciación Acumulada*": 21,
    "Activos Intangibles": 22,
    "Otros activos no financieros (no corriente)": 23,
    "Obligaciones financieras (corriente)": 26,
    "Cuentas por Pagar Comerciales": 27,
    "Cuentas por Pagar a Entidades Relacionadas": 28,
    "Otras Cuentas por Pagar": 29,
    "Otros pasivos (corriente)": 30,
    "Obligaciones financieras (no corriente)": 32,
    "Cuentas por Pagar Comerciales y Otras Cuentas por Pagar": 33,
    "Otros pasivos (no corriente)": 34,
    "Capital": 37,
    "Reserva legal y otras reservas": 38,
    "Resultados Acumulados": 39,
    "Otros": 40,
}

RESULTS_ROWS = {
    "Ingresos de Actividades Ordinarias": 46,
    "Costo de Ventas": 47,
    "Gastos de Ventas y Distribución": 49,
    "Depreciación*": 50,
    "Gastos de Administración": 51,
    "Otros ingresos (gastos) netos": 52,
    "Ingresos financieros": 54,
    "Gastos financieros": 55,
    "Diferencia en cambio, neta": 56,
    "Impuesto a la renta": 58,
}


def _normalize(label: str) -> str:
    return re.sub(r"\s+", " ", str(label).strip().casefold())


def _row_map(sheet, start: int, end: int) -> dict[str, int]:
    result = {}
    for row in range(start, end + 1):
        label = sheet.cell(row=row, column=2).value
        if label:
            result[_normalize(label)] = row
    return result


def _year_columns(first_year: int, last_year: int) -> dict[int, int]:
    if last_year < first_year or last_year - first_year + 1 > 10:
        raise ValueError("La plantilla permite entre 1 y 10 años históricos")
    return {
        year: 3 + index
        for index, year in enumerate(range(first_year, last_year + 1))
    }


def fill_valora_template(template_bytes: bytes, extracted: dict) -> bytes:
    """Return a filled copy of the template, preserving formulas and formatting."""
    workbook = load_workbook(BytesIO(template_bytes))
    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"La plantilla no contiene la hoja {SHEET_NAME!r}")
    sheet = workbook[SHEET_NAME]

    metadata = extracted.get("metadata") or {}
    periods = {int(year) for year in metadata.get("periodos", []) if str(year).isdigit()}
    for table_key in ("balance_table", "results_table"):
        table = extracted.get(table_key) or {}
        periods.update(int(year) for year in table.get("years", []) if str(year).isdigit())
    periods = sorted(periods)
    if not periods:
        raise ValueError("La extracción no contiene períodos históricos")

    first_year, last_year = periods[0], periods[-1]
    columns = _year_columns(first_year, last_year)
    sheet[FIRST_YEAR_CELL] = first_year
    sheet[LAST_YEAR_CELL] = last_year
    shares = extracted.get("number_of_shares")
    if isinstance(shares, dict):
        shares = shares.get("value")
    if shares not in (None, ""):
        sheet["C5"] = shares

    balance_rows = _row_map(sheet, 11, 42)
    result_rows = _row_map(sheet, 46, 59)

    def write_section(section: str, configured_rows: dict[str, int], discovered_rows: dict[str, int]) -> None:
        source = (extracted.get(section) or {})
        for account, values in source.items():
            # Prefer explicit rows because the template repeats labels such as
            # "Otros pasivos" in current and non-current sections.
            row = configured_rows.get(account)
            if row is None:
                row = discovered_rows.get(_normalize(account))
            if row is None or not isinstance(values, dict):
                continue
            for period, value in values.items():
                try:
                    column = columns[int(period)]
                except (KeyError, TypeError, ValueError):
                    continue
                if value is not None:
                    sheet.cell(row=row, column=column).value = value

    def table_to_source(table: dict | None) -> dict:
        source = {}
        duplicate_rows = {
            "obligaciones financieras": ("Obligaciones financieras (corriente)", "Obligaciones financieras (no corriente)"),
            "otros pasivos": ("Otros pasivos (corriente)", "Otros pasivos (no corriente)"),
        }
        occurrences = {}
        if not table:
            return source
        years = [str(year) for year in table.get("years", [])]
        for item in table.get("rows", []):
            values = item.get("values", [])
            label = item.get("label", "")
            normalized = _normalize(label)
            index = occurrences.get(normalized, 0)
            occurrences[normalized] = index + 1
            duplicate_labels = duplicate_rows.get(normalized)
            target_label = (
                duplicate_labels[min(index, len(duplicate_labels) - 1)]
                if duplicate_labels
                else label
            )
            source[target_label] = {
                year: values[index] if index < len(values) else None
                for index, year in enumerate(years)
            }
        return source

    extracted = {
        **extracted,
        "balance_general": extracted.get("balance_general") or table_to_source(extracted.get("balance_table")),
        "estado_resultados": extracted.get("estado_resultados") or table_to_source(extracted.get("results_table")),
    }
    write_section("balance_general", BALANCE_ROWS, balance_rows)
    write_section("estado_resultados", RESULTS_ROWS, result_rows)

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
