"""
Servicio para extraer Template Codes (códigos con formato $$XXXXX$$) 
de archivos Excel (Master Templates).

Los códigos se extraen de hojas específicas:
- "Plantilla Usuario" -> type: VALORA
- "WACC" -> type: KAPITAL
"""

import re
import logging
from typing import List, Dict, Literal, Optional
from io import BytesIO
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


logger = logging.getLogger(__name__)

TemplateType = Literal["valora", "kapital"]
SheetMapping = Dict[str, TemplateType]

# Mapeo de hojas a tipos de template
DEFAULT_SHEET_MAPPING: SheetMapping = {
    "Plantilla Usuario": "valora",
    "WACC": "kapital",
}

# Patrón para detectar códigos de template
CODE_PATTERN = re.compile(r"\$\$([A-Z0-9]+)\$\$")


class TemplateCodeExtractor:
    """Extrae Template Codes de archivos Excel."""

    def __init__(self, sheet_mapping: Optional[SheetMapping] = None):
        """
        Inicializa el extractor.
        
        Args:
            sheet_mapping: Mapeo personalizado de hojas a tipos.
                          Si es None, usa DEFAULT_SHEET_MAPPING.
        """
        self.sheet_mapping = sheet_mapping or DEFAULT_SHEET_MAPPING

    def extract_from_bytes(
        self,
        file_content: bytes,
        extract_images: bool = False,
    ) -> Dict[str, List[Dict]]:
        """
        Extrae códigos de un contenido de archivo Excel.
        
        Args:
            file_content: Contenido del archivo Excel en bytes
            extract_images: Si True, intenta extraer imágenes con códigos
        
        Returns:
            Dict con estructura:
            {
                "valora": [...codes...],
                "kapital": [...codes...],
                "extracted_count": int,
                "processed_sheets": list
            }
        """
        try:
            # Cargar con data_only=True para obtener valores calculados
            workbook = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
            return self._extract_from_workbook(workbook, extract_images)
        except Exception as e:
            logger.error(f"Error extracting codes from Excel: {e}")
            return {
                "valora": [],
                "kapital": [],
                "extracted_count": 0,
                "processed_sheets": [],
                "error": str(e)
            }

    def _extract_from_workbook(
        self,
        workbook: openpyxl.Workbook,
        extract_images: bool = False,
    ) -> Dict[str, List[Dict]]:
        """Extrae códigos del workbook."""
        result = {
            "valora": [],
            "kapital": [],
            "extracted_count": 0,
            "processed_sheets": [],
        }

        # Procesar solo las hojas que están en el mapeo
        for sheet_name in workbook.sheetnames:
            if sheet_name not in self.sheet_mapping:
                continue

            template_type = self.sheet_mapping[sheet_name]
            worksheet = workbook[sheet_name]
            
            codes = self._extract_from_sheet(worksheet, sheet_name, template_type)
            result[template_type].extend(codes)
            result["processed_sheets"].append({
                "name": sheet_name,
                "type": template_type,
                "codes_count": len(codes)
            })

            # Extraer de imágenes si es necesario
            if extract_images:
                image_codes = self._extract_from_images(worksheet, template_type)
                result[template_type].extend(image_codes)

        result["extracted_count"] = len(result["valora"]) + len(result["kapital"])
        return result

    def _sanitize_cell_value(self, cell_value) -> Optional[str]:
        """
        Sanitiza el valor de una celda.
        Rechaza objetos openpyxl y retorna solo strings limpios.
        
        Args:
            cell_value: Valor bruto de la celda
            
        Returns:
            String limpio o None si es inválido
        """
        if cell_value is None:
            return None
            
        # Si ya es un string, retornarlo limpio
        if isinstance(cell_value, str):
            return cell_value.strip()
        
        # Si es un número, convertirlo a string
        if isinstance(cell_value, (int, float, bool)):
            return str(cell_value).strip()
        
        # Convertir a string y verificar que no sea un objeto openpyxl
        cell_str = str(cell_value)
        
        # Rechazar si contiene representaciones de objetos openpyxl
        if "<openpyxl" in cell_str or "object at 0x" in cell_str:
            logger.warning(f"Rejected openpyxl object in cell: {cell_str[:100]}")
            return None
        
        # Rechazar si contiene "<" y ">" que indica objeto extraño
        if cell_str.startswith("<") and "object" in cell_str:
            logger.warning(f"Rejected suspicious object in cell: {cell_str[:100]}")
            return None
        
        return cell_str.strip() if cell_str else None

    def _extract_from_sheet(
        self,
        worksheet: Worksheet,
        sheet_name: str,
        template_type: TemplateType,
    ) -> List[Dict]:
        """Extrae códigos de todas las celdas de una hoja."""
        codes = []
        seen_codes = set()  # Para evitar duplicados

        # Iterar sobre todas las celdas con valor
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue

                # Sanitizar el valor de la celda (rechaza openpyxl objects)
                cell_value = self._sanitize_cell_value(cell.value)
                
                if not cell_value:
                    continue

                # Buscar códigos en la celda
                matches = CODE_PATTERN.findall(cell_value)
                for code in matches:
                    code_formatted = f"$${code}$$"

                    # Evitar duplicados
                    if code_formatted in seen_codes:
                        continue
                    seen_codes.add(code_formatted)

                    codes.append({
                        "code": code_formatted,
                        "nombre": self._get_cell_name(cell),
                        "hoja": sheet_name,
                        "type": template_type,
                    })

        return codes

    def _get_cell_name(self, cell) -> str:
        """
        Obtiene el nombre de la celda mirando SOLO a la DERECHA (columna F).
        
        Patrón esperado:
        - Columna E: $$CODIGO$$
        - Columna F: Nombre del código (la ÚNICA columna de nombres)
        
        Si la columna F está vacía, retorna "NA".
        Las imágenes/gráficos siguen el mismo patrón con su nombre en la celda de al lado.
        """
        # Obtener nombre de F (col+1) - LA ÚNICA FUENTE DE NOMBRES
        if cell.column < cell.parent.max_column:
            next_cell = cell.parent.cell(cell.row, cell.column + 1)
            next_value = self._sanitize_cell_value(next_cell.value)
            if next_value and "$$" not in next_value:
                return next_value
        
        # Si F está vacío, retornar "NA"
        return "NA"

    def get_statistics(self, extraction_result: Dict) -> Dict:
        """Obtiene estadísticas de la extracción."""
        return {
            "total_codes": extraction_result.get("extracted_count", 0),
            "valora_codes": len(extraction_result.get("valora", [])),
            "kapital_codes": len(extraction_result.get("kapital", [])),
            "sheets_processed": len(extraction_result.get("processed_sheets", [])),
            "sheets": extraction_result.get("processed_sheets", [])
        }


# Singleton para reutilizar
_extractor_instance: Optional[TemplateCodeExtractor] = None


def get_template_code_extractor() -> TemplateCodeExtractor:
    """Obtiene la instancia singleton del extractor."""
    global _extractor_instance
    if _extractor_instance is None:
        _extractor_instance = TemplateCodeExtractor()
    return _extractor_instance
